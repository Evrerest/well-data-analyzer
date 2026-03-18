from __future__ import annotations

from math import ceil, floor, log10
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from plotly.colors import qualitative


DEFAULT_SAMPLE_PATH = Path(r"C:\Users\ekost\Downloads\Viking Core Data.xlsx")
KNOWN_HEADERS = {
    1: "Area",
    2: "Well",
    3: "Core",
    4: "Code",
    5: "Depth (M)",
    6: "Length (M)",
    7: "K-Max (mD)",
    8: "k90 (mD)",
    9: "k-vert (mD)",
    10: "Porosity (%)",
    11: "Grn Den (Kg/m3)",
    12: "Blk Den (Kg/m3)", 
    13: "Blk Mass Oil (%)",
    14: "Blk Mass Wtr (%)",
    15: "Blk Mass Snd (%)",
    16: "Por Vol Oil (%)",
    17: "Por Vol Wtr (%)",
    18: "Blk Vol Oil (%)",
    19: "Blk Vol Wtr (%)",
    20: "Grn Mass Oil (%)",
    21: "Grn Mass Wtr (%)",
    22: "Lithology",
    23: "Kmax-H",
    24: "Phi-H",
    25: "Formation",
}
DEFAULT_X = "Porosity (%)"
DEFAULT_Y = "K-Max (mD)"
AREA_FALLBACK = "Unassigned Area"
DEFAULT_X_START = 0.0
DEFAULT_X_END = 0.2
DEFAULT_X_MAJOR = 0.02
DEFAULT_X_MINOR = 0.005
DEFAULT_Y_START = 0.01
DEFAULT_Y_END = 10000.0
DEFAULT_AREA_COLORS = [
    "#858484",
    "#c85a5a",
    "#3980C3",
    "#4f8a54",
    "#e38b8b",
    "#7b5aa6",
]


def set_highlighted_well(well: str | None) -> None:
    st.session_state["highlighted_well"] = well


def start_line_selection() -> None:
    st.session_state["line_selection_mode"] = True
    st.session_state["selected_line_points"] = []


def clear_line_selection() -> None:
    st.session_state["line_selection_mode"] = False
    st.session_state["selected_line_points"] = []


def resolve_default_workbook_path() -> str:
    candidate_paths = []

    bundle_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidate_paths.append(bundle_dir / "sample_data" / "Viking Core Data.xlsx")
    candidate_paths.append(DEFAULT_SAMPLE_PATH)

    for candidate in candidate_paths:
        if candidate.exists():
            return str(candidate)
    return ""


def sanitize_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return value


def build_headers(worksheet) -> list[str]:
    headers: list[str] = []
    for column in range(1, worksheet.max_column + 1):
        if column in KNOWN_HEADERS:
            headers.append(KNOWN_HEADERS[column])
            continue

        row2 = sanitize_value(worksheet.cell(2, column).value)
        if row2 is not None:
            headers.append(str(row2))
            continue

        headers.append(f"Column_{column}")
    return headers


@st.cache_data(show_spinner=False)
def load_workbook_data(workbook_path: str) -> tuple[str, pd.DataFrame]:
    path = Path(workbook_path)
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    title = str(worksheet["A1"].value or path.stem)
    headers = build_headers(worksheet)

    current_area: str | None = None
    current_well: str | None = None
    records: list[dict[str, Any]] = []

    for row_number in range(3, worksheet.max_row + 1):
        row_values = [sanitize_value(worksheet.cell(row_number, col).value) for col in range(1, worksheet.max_column + 1)]
        if not any(value is not None for value in row_values):
            continue

        if row_values[0] is not None:
            current_area = str(row_values[0])
        if row_values[1] is not None:
            current_well = str(row_values[1])

        record = {header: value for header, value in zip(headers, row_values)}
        record["Area"] = current_area or AREA_FALLBACK
        record["Well"] = current_well or "Unknown Well"
        records.append(record)

    dataframe = pd.DataFrame(records)
    for column in dataframe.columns:
        converted = pd.to_numeric(dataframe[column], errors="coerce")
        non_null_original = dataframe[column].notna().sum()
        non_null_converted = converted.notna().sum()
        if non_null_original > 0 and non_null_original == non_null_converted:
            dataframe[column] = converted

    if "Area" not in dataframe:
        dataframe["Area"] = AREA_FALLBACK
    if "Well" not in dataframe:
        dataframe["Well"] = "Unknown Well"

    dataframe["LegendLabel"] = dataframe["Area"].astype(str) + " | " + dataframe["Well"].astype(str)
    return title, dataframe


def build_color_map(labels: list[str], overrides: dict[str, str] | None = None) -> dict[str, str]:
    palette = DEFAULT_AREA_COLORS + qualitative.Safe + qualitative.Bold + qualitative.Dark24
    color_map = {label: palette[index % len(palette)] for index, label in enumerate(labels)}
    if overrides:
        color_map.update({label: color for label, color in overrides.items() if color})
    return color_map


def nice_linear_tick(raw_step: float | None) -> float | None:
    if raw_step is None or raw_step <= 0:
        return None

    magnitude = 10 ** int(log10(raw_step))
    normalized = raw_step / magnitude
    if normalized <= 1:
        nice = 1
    elif normalized <= 2:
        nice = 2
    elif normalized <= 5:
        nice = 5
    else:
        nice = 10
    return nice * magnitude


def default_tick_step(series: pd.Series) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return None
    data_min = float(numeric.min())
    data_max = float(numeric.max())
    if data_max <= data_min:
        return None
    return nice_linear_tick((data_max - data_min) / 8)


def default_minor_step(major_step: float | None) -> float | None:
    if major_step is None or major_step <= 0:
        return None
    return major_step / 5


def build_log_tick_values(start: float, end: float, minor_per_decade: int) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []

    first_decade = int(log10(start))
    last_decade = int(log10(end)) + 1
    ticks: list[float] = []

    for exponent in range(first_decade, last_decade + 1):
        decade_base = 10**exponent
        for multiplier in range(1, minor_per_decade + 1):
            tick_value = multiplier * decade_base
            if start <= tick_value <= end:
                ticks.append(float(tick_value))
    return ticks

def build_linear_tick_values(start: float, end: float, step: float) -> list[float]:
    if step <= 0 or end <= start:
        return []

    ticks: list[float] = []
    first_index = ceil(start / step)
    current = first_index * step
    limit = end + (step * 1e-9)
    while current <= limit:
        ticks.append(round(current, 12))
        current += step
    return ticks


def build_log_major_ticks(start: float, end: float, major_factor: float) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []
    if major_factor <= 1:
        major_factor = 10

    ticks: list[float] = []
    current_power = ceil(log10(start) / log10(major_factor))
    current = major_factor**current_power
    while current <= end:
        ticks.append(round(float(current), 12))
        current *= major_factor
    return ticks


def build_log_minor_ticks(start: float, end: float, major_factor: float, subdivisions: int) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []
    if major_factor <= 1 or subdivisions <= 1:
        return []

    ticks: list[float] = []
    current_power = floor(log10(start) / log10(major_factor))
    current_major = major_factor**current_power
    while current_major <= end:
        for index in range(2, subdivisions + 1):
            tick = current_major * index
            if start <= tick <= end and tick < current_major * major_factor:
                ticks.append(round(float(tick), 12))
        current_major *= major_factor
    return ticks


def build_picker_axis_values(start: float, end: float, scale: str, count: int = 45) -> list[float]:
    if end <= start or count < 2:
        return []

    if scale == "Logarithmic":
        log_start = log10(start)
        log_end = log10(end)
        return [10 ** (log_start + ((log_end - log_start) * index / (count - 1))) for index in range(count)]

    return [start + ((end - start) * index / (count - 1)) for index in range(count)]


def resolve_axis_bounds(
    series: pd.Series,
    scale: str,
    start: float | None,
    end: float | None,
) -> tuple[float | None, float | None]:
    if start is not None and end is not None and end > start:
        return float(start), float(end)

    if scale == "Logarithmic":
        lower = default_log_axis_value(series, "min")
        upper = default_log_axis_value(series, "max")
    else:
        lower = default_axis_value(series, "min")
        upper = default_axis_value(series, "max")

    if lower is None or upper is None or upper <= lower:
        return None, None
    return float(lower), float(upper)


def apply_axis_settings(
    figure: go.Figure,
    axis_name: str,
    title: str,
    scale: str,
    start: float | None,
    end: float | None,
    major_tick: float | None,
    minor_tick: float | None,
) -> None:
    axis_type = "log" if scale == "Logarithmic" else "linear"
    update_method = figure.update_xaxes if axis_name == "x" else figure.update_yaxes

    axis_args: dict[str, Any] = {
        "title_text": "",
        "type": axis_type,
        "showgrid": True,
        "gridcolor": "rgba(70, 70, 70, 0.38)",
        "zeroline": False,
        "tickformat": "g",
        "title_font": {"size": 20, "color": "#111111"},
        "tickfont": {"size": 15, "color": "#222222"},
    }

    if start is not None and end is not None and end > start:
        if axis_type == "log":
            if start > 0 and end > 0:
                axis_args["range"] = [log10(start), log10(end)]
        else:
            axis_args["range"] = [start, end]

    if axis_type == "linear":
        if start is not None and end is not None and major_tick and major_tick > 0:
            major_ticks = build_linear_tick_values(start, end, major_tick)
            axis_args["tickmode"] = "array"
            axis_args["tickvals"] = major_ticks
            axis_args["ticktext"] = [f"{tick:g}" for tick in major_ticks]

        if start is not None and end is not None and minor_tick and minor_tick > 0:
            minor_ticks = build_linear_tick_values(start, end, minor_tick)
            major_tick_set = set(axis_args.get("tickvals", []))
            minor_only = [tick for tick in minor_ticks if tick not in major_tick_set]
            axis_args["minor"] = {
                "showgrid": True,
                "gridcolor": "rgba(120, 120, 120, 0.14)",
                "tickmode": "array",
                "tickvals": minor_only,
            }
    else:
        if start is not None and end is not None and start > 0 and end > start:
            major_ticks = build_log_major_ticks(start, end, major_tick or 10)
            if major_ticks:
                axis_args["tickmode"] = "array"
                axis_args["tickvals"] = major_ticks
                axis_args["ticktext"] = [f"{tick:g}" for tick in major_ticks]

            if minor_tick and minor_tick > 1:
                subdivisions = max(int(minor_tick), 1)
                axis_args["minor"] = {
                    "showgrid": True,
                    "gridcolor": "rgba(120, 120, 120, 0.14)",
                    "tickmode": "array",
                    "tickvals": build_log_minor_ticks(start, end, major_tick or 10, subdivisions),
                }

    update_method(**axis_args)


def build_figure(
    dataframe: pd.DataFrame,
    chart_title: str,
    area_color_map: dict[str, str],
    highlighted_well: str | None,
    line_selection_mode: bool,
    x_column: str,
    y_column: str,
    x_scale: str,
    y_scale: str,
    x_start: float | None,
    x_end: float | None,
    x_major: float | None,
    x_minor: float | None,
    y_start: float | None,
    y_end: float | None,
    y_major: float | None,
    y_minor: float | None,
    selected_line_points: list[dict[str, float]] | None,
) -> go.Figure:
    plot_data = dataframe[[x_column, y_column, "Area", "Well", "LegendLabel"]].copy()
    plot_data = plot_data.dropna(subset=[x_column, y_column])
    plot_data = plot_data[(plot_data[x_column] != "") & (plot_data[y_column] != "")]

    if x_scale == "Logarithmic":
        plot_data = plot_data[pd.to_numeric(plot_data[x_column], errors="coerce") > 0]
    if y_scale == "Logarithmic":
        plot_data = plot_data[pd.to_numeric(plot_data[y_column], errors="coerce") > 0]

    if plot_data.empty:
        figure = go.Figure()
        figure.update_layout(
            height=720,
            paper_bgcolor="#e7e4dc",
            plot_bgcolor="#ffffff",
            annotations=[
                {
                    "text": "No rows match the current axis settings.",
                    "xref": "paper",
                    "yref": "paper",
                    "x": 0.5,
                    "y": 0.5,
                    "showarrow": False,
                    "font": {"size": 18},
                }
            ],
        )
        return figure

    figure = px.scatter(
        plot_data,
        x=x_column,
        y=y_column,
        color="Area",
        color_discrete_map=area_color_map,
        hover_data={"Area": True, "Well": True},
        custom_data=["Well"],
        labels={x_column: x_column, y_column: y_column, "Area": "Area"},
    )
    figure.update_traces(marker={"size": 7.5, "line": {"width": 0.5, "color": "#1f2933"}})

    if highlighted_well:
        highlighted_data = plot_data[plot_data["Well"] == highlighted_well]
        if not highlighted_data.empty:
            figure.add_trace(
                go.Scatter(
                    x=highlighted_data[x_column],
                    y=highlighted_data[y_column],
                    mode="markers",
                    name=f"Selected: {highlighted_well}",
                    showlegend=False,
                    hoverinfo="skip",
                    marker={
                        "size": 10,
                        "color": "rgba(0, 0, 0, 0)",
                        "line": {"width": 3, "color": "#ffeb3b"},
                    },
                )
            )

    if selected_line_points and len(selected_line_points) == 2:
        figure.add_trace(
            go.Scatter(
                x=[selected_line_points[0]["x"], selected_line_points[1]["x"]],
                y=[selected_line_points[0]["y"], selected_line_points[1]["y"]],
                mode="lines+markers",
                name="Selected Line",
                showlegend=False,
                hoverinfo="skip",
                line={"color": "#111111", "width": 2},
                marker={
                    "size": 10,
                    "color": "#ffeb3b",
                    "line": {"width": 2, "color": "#111111"},
                },
            )
        )

    x_picker_start, x_picker_end = resolve_axis_bounds(dataframe[x_column], x_scale, x_start, x_end)
    y_picker_start, y_picker_end = resolve_axis_bounds(dataframe[y_column], y_scale, y_start, y_end)
    if line_selection_mode and x_picker_start is not None and x_picker_end is not None and y_picker_start is not None and y_picker_end is not None:
        x_picker_values = build_picker_axis_values(x_picker_start, x_picker_end, x_scale)
        y_picker_values = build_picker_axis_values(y_picker_start, y_picker_end, y_scale)
        picker_x: list[float] = []
        picker_y: list[float] = []
        picker_customdata: list[str] = []
        for x_value in x_picker_values:
            for y_value in y_picker_values:
                picker_x.append(x_value)
                picker_y.append(y_value)
                picker_customdata.append(f"picker|{x_value:.12g}|{y_value:.12g}")

        figure.add_trace(
            go.Scatter(
                x=picker_x,
                y=picker_y,
                mode="markers",
                name="Line Picker",
                showlegend=False,
                hoverinfo="none",
                customdata=picker_customdata,
                marker={
                    "size": 11,
                    "color": "rgba(0, 0, 0, 0.01)",
                    "line": {"width": 0},
                },
            )
        )

    figure.update_layout(
        height=720,
        margin={"l": 90, "r": 90, "t": 90, "b": 90},
        legend_title_text="Legend",
        clickmode="event+select",
        paper_bgcolor="#e7e4dc",
        plot_bgcolor="#ffffff",
        annotations=[
            {
                "text": chart_title,
                "xref": "paper",
                "yref": "paper",
                "x": 0.5,
                "y": 1.09,
                "xanchor": "center",
                "yanchor": "middle",
                "showarrow": False,
                "font": {"size": 24, "color": "#111111"},
            },
            {
                "text": x_column,
                "xref": "paper",
                "yref": "paper",
                "x": 0.5,
                "y": -0.12,
                "xanchor": "center",
                "yanchor": "middle",
                "showarrow": False,
                "font": {"size": 20, "color": "#111111"},
            },
            {
                "text": y_column,
                "xref": "paper",
                "yref": "paper",
                "x": -0.085,
                "y": 0.5,
                "xanchor": "center",
                "yanchor": "middle",
                "textangle": -90,
                "showarrow": False,
                "font": {"size": 20, "color": "#111111"},
            },
        ],
        shapes=[
            {
                "type": "rect",
                "xref": "paper",
                "yref": "paper",
                "x0": 0,
                "y0": 0,
                "x1": 1,
                "y1": 1,
                "line": {"color": "rgba(0, 0, 0, 0.25)", "width": 1},
                "fillcolor": "rgba(0, 0, 0, 0)",
                "layer": "below",
            }
        ],
        legend={"x": 1.02, "y": 0.84, "font": {"size": 19}, "title_font": {"size": 17}},
    )

    apply_axis_settings(figure, "x", x_column, x_scale, x_start, x_end, x_major, x_minor)
    apply_axis_settings(figure, "y", y_column, y_scale, y_start, y_end, y_major, y_minor)
    figure.update_xaxes(
        tickfont={"size": 15, "color": "#222222"},
    )
    figure.update_yaxes(
        tickfont={"size": 15, "color": "#222222"},
        ticklabelstandoff=10,
    )
    return figure


def extract_selected_well(selection_event: Any) -> str | None:
    if not selection_event:
        return None

    points = selection_event.get("selection", {}).get("points", [])
    if not points:
        return None

    first_point = points[0]
    customdata = first_point.get("customdata")
    if isinstance(customdata, list) and customdata:
        return str(customdata[0])
    if isinstance(customdata, str):
        return customdata
    return None


def extract_selected_points(selection_event: Any) -> list[dict[str, float]]:
    if not selection_event:
        return []

    points = selection_event.get("selection", {}).get("points", [])
    selected_points: list[dict[str, float]] = []
    for point in points:
        customdata = point.get("customdata")
        if isinstance(customdata, str) and customdata.startswith("picker|"):
            _, x_value, y_value = customdata.split("|", maxsplit=2)
            selected_points.append({"x": float(x_value), "y": float(y_value)})
            continue

        x_value = point.get("x")
        y_value = point.get("y")
        if x_value is None or y_value is None:
            continue
        selected_points.append({"x": float(x_value), "y": float(y_value)})
    return selected_points


def has_chart_selection(selection_event: Any) -> bool:
    if not selection_event:
        return False
    points = selection_event.get("selection", {}).get("points", [])
    return bool(points)


def append_line_point(selection_event: Any) -> None:
    existing_points = list(st.session_state.get("selected_line_points", []))
    incoming_points = extract_selected_points(selection_event)
    if not incoming_points:
        return

    if len(incoming_points) >= 2:
        st.session_state["selected_line_points"] = incoming_points[:2]
        st.session_state["line_selection_mode"] = False
        return

    point = incoming_points[0]
    if existing_points:
        previous = existing_points[-1]
        if previous["x"] == point["x"] and previous["y"] == point["y"]:
            return

    existing_points.append(point)
    st.session_state["selected_line_points"] = existing_points[:2]
    if len(st.session_state["selected_line_points"]) >= 2:
        st.session_state["line_selection_mode"] = False


def handle_chart_selection(chart_key: str) -> None:
    selection_event = st.session_state.get(chart_key)
    if not has_chart_selection(selection_event):
        return

    if st.session_state.get("line_selection_mode"):
        append_line_point(selection_event)
        return

    selected_well = extract_selected_well(selection_event)
    if selected_well is not None:
        st.session_state["highlighted_well"] = selected_well


def format_equation_value(value: float) -> str:
    if value == 0:
        return "0"
    if abs(value) >= 1000 or abs(value) < 0.001:
        return f"{value:.4e}"
    return f"{value:.6g}"


def build_line_equation(
    line_points: list[dict[str, float]] | None,
    x_scale: str,
    y_scale: str,
) -> str | None:
    if not line_points or len(line_points) != 2:
        return None

    x1 = line_points[0]["x"]
    y1 = line_points[0]["y"]
    x2 = line_points[1]["x"]
    y2 = line_points[1]["y"]

    if x_scale == "Logarithmic" and (x1 <= 0 or x2 <= 0):
        return None
    if y_scale == "Logarithmic" and (y1 <= 0 or y2 <= 0):
        return None

    x1_transformed = log10(x1) if x_scale == "Logarithmic" else x1
    x2_transformed = log10(x2) if x_scale == "Logarithmic" else x2
    y1_transformed = log10(y1) if y_scale == "Logarithmic" else y1
    y2_transformed = log10(y2) if y_scale == "Logarithmic" else y2

    if x1_transformed == x2_transformed:
        return f"x = {format_equation_value(x1)}"

    slope = (y2_transformed - y1_transformed) / (x2_transformed - x1_transformed)
    intercept = y1_transformed - (slope * x1_transformed)
    intercept_sign = "+" if intercept >= 0 else "-"
    signed_intercept = f"{intercept_sign} {format_equation_value(abs(intercept))}"

    if x_scale == "Linear" and y_scale == "Linear":
        return f"y = {format_equation_value(slope)}x {signed_intercept}"

    if x_scale == "Linear" and y_scale == "Logarithmic":
        return (
            f"log10(y) = {format_equation_value(slope)}x {signed_intercept}    "
            f"y = 10^({format_equation_value(slope)}x {signed_intercept})"
        )

    if x_scale == "Logarithmic" and y_scale == "Linear":
        return f"y = {format_equation_value(slope)}log10(x) {signed_intercept}"

    coefficient = 10**intercept
    return (
        f"log10(y) = {format_equation_value(slope)}log10(x) {signed_intercept}    "
        f"y = {format_equation_value(coefficient)}x^{format_equation_value(slope)}"
    )


def default_axis_value(series: pd.Series, which: str) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return None
    return float(numeric.min() if which == "min" else numeric.max())


def default_log_axis_value(series: pd.Series, which: str) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce")
    numeric = numeric[numeric > 0].dropna()
    if numeric.empty:
        return None
    return float(numeric.min() if which == "min" else numeric.max())


def grouped_legend(dataframe: pd.DataFrame) -> dict[str, list[str]]:
    grouped = (
        dataframe[["Area", "Well"]]
        .drop_duplicates()
        .sort_values(["Area", "Well"])
        .groupby("Area")["Well"]
        .apply(list)
        .to_dict()
    )
    return grouped


def axis_input_block(axis_id: str, series: pd.Series, scale: str) -> tuple[float, float, float, float]:
    if scale == "Logarithmic":
        min_value = DEFAULT_Y_START if axis_id == "y" else (default_log_axis_value(series, "min") or 0.1)
        max_value = DEFAULT_Y_END if axis_id == "y" else (default_log_axis_value(series, "max") or 10.0)
        major_default = 10.0
        major_label = "Major Division Factor"
        major_help = "Use 10 for decade spacing: 0.01, 0.1, 1, 10, 100."
    else:
        if axis_id == "x":
            min_value = DEFAULT_X_START
            max_value = DEFAULT_X_END
            major_default = DEFAULT_X_MAJOR
            minor_default = DEFAULT_X_MINOR
        else:
            min_value = default_axis_value(series, "min") or 0.0
            max_value = default_axis_value(series, "max") or 1.0
            major_default = default_tick_step(series) or 1.0
            minor_default = default_minor_step(major_default) or 0.2
        major_label = "Major Division Spacing"
        minor_label = "Minor Division Spacing"
        major_help = "Exact spacing between major divisions."
        minor_help = "Exact spacing between lighter minor divisions."

    start = st.number_input(
        "Start",
        value=min_value,
        format="%.6f",
        key=f"{axis_id}_start_{series.name}",
    )
    end = st.number_input(
        "End",
        value=max_value,
        format="%.6f",
        key=f"{axis_id}_end_{series.name}",
    )
    major = st.number_input(
        major_label,
        min_value=float(0.0),
        value=float(major_default),
        format="%.6f",
        key=f"{axis_id}_major_{series.name}",
        help=major_help,
    )
    if scale == "Logarithmic":
        show_minor = st.toggle(
            "Show Minor Divisions",
            value=True,
            key=f"{axis_id}_minor_toggle_{series.name}",
            help="Off shows no minor log divisions. On shows 2 through 9 within each decade.",
        )
        minor = 9.0 if show_minor else 0.0
    else:
        minor = st.number_input(
            minor_label,
            min_value=float(0.0),
            value=float(minor_default),
            format="%.6f",
            key=f"{axis_id}_minor_{series.name}",
            help=minor_help,
        )
    return start, end, major, minor


def validate_axis_settings(axis_label: str, scale: str, start: float, end: float, major: float) -> list[str]:
    errors: list[str] = []

    if end <= start:
        errors.append(f"{axis_label}: End must be greater than Start.")

    if scale == "Logarithmic":
        if start <= 0 or end <= 0:
            errors.append(f"{axis_label}: Logarithmic axes require Start and End to be greater than 0.")
        if major <= 1:
            errors.append(f"{axis_label}: Major Division Factor must be greater than 1. Use 10 for decades.")
    else:
        if major <= 0:
            errors.append(f"{axis_label}: Major Division Spacing must be greater than 0.")

    return errors


st.set_page_config(page_title="Well Data Analyzer", layout="wide")

st.markdown(
    """
    <style>
    div[data-testid="stButton"] > button[kind="primary"] {
        background-color: #ffeb3b;
        border-color: #d4c019;
        color: #111111;
    }
    div[data-testid="stButton"] > button[kind="primary"]:hover {
        background-color: #fff176;
        border-color: #d4c019;
        color: #111111;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Well Data Analyzer")
st.caption("Interactive cross-plotting for core and well data.")

st.session_state.setdefault("line_selection_mode", False)
st.session_state.setdefault("selected_line_points", [])

with st.sidebar:
    st.header("Workbook")
    uploaded_file = st.file_uploader("Upload an Excel workbook", type=["xlsx"])
    workbook_path = st.text_input("Or enter a local workbook path", value=resolve_default_workbook_path())

if uploaded_file is not None:
    upload_dir = Path(".streamlit_uploads")
    upload_dir.mkdir(exist_ok=True)
    uploaded_path = upload_dir / uploaded_file.name
    uploaded_path.write_bytes(uploaded_file.getbuffer())
    selected_path = str(uploaded_path.resolve())
else:
    selected_path = workbook_path

path_obj = Path(selected_path)
if not selected_path:
    st.info("Upload a workbook or enter a workbook path to begin.")
    st.stop()

if not path_obj.exists():
    st.error(f"Workbook not found: {path_obj}")
    st.stop()

try:
    workbook_title, data = load_workbook_data(str(path_obj))
except Exception as exc:
    st.exception(exc)
    st.stop()

st.subheader(workbook_title)

numeric_columns = [column for column in data.columns if pd.api.types.is_numeric_dtype(data[column])]
if not numeric_columns:
    st.error("No numeric columns were found in the workbook.")
    st.stop()

default_x_index = numeric_columns.index(DEFAULT_X) if DEFAULT_X in numeric_columns else 0
default_y_index = numeric_columns.index(DEFAULT_Y) if DEFAULT_Y in numeric_columns else min(1, len(numeric_columns) - 1)

legend_groups = grouped_legend(data)
legend_areas = data["Area"].drop_duplicates().tolist()
base_area_color_map = build_color_map(legend_areas)
area_color_overrides: dict[str, str] = {}

left_col, right_col = st.columns([1, 3], gap="large")

with right_col:
    graph_tab, settings_tab, preview_tab = st.tabs(["Graph", "Settings", "Data Preview"])

    with settings_tab:
        selector_col, axis_col = st.columns([1, 2], gap="large")

        with selector_col:
            x_column = st.selectbox("X Axis", numeric_columns, index=default_x_index)
            y_column = st.selectbox("Y Axis", numeric_columns, index=default_y_index)

        with axis_col:
            x_settings, y_settings = st.columns(2, gap="large")

            with x_settings:
                st.markdown("#### X Axis Settings")
                x_scale = st.radio("Scale", ["Linear", "Logarithmic"], key="x_scale", index=0)
                x_start, x_end, x_major, x_minor = axis_input_block("x", data[x_column], x_scale)

            with y_settings:
                st.markdown("#### Y Axis Settings")
                y_scale = st.radio("Scale", ["Linear", "Logarithmic"], key="y_scale", index=1)
                y_start, y_end, y_major, y_minor = axis_input_block("y", data[y_column], y_scale)

        st.caption(
            "Linear axes use exact spacing values. "
            "Logarithmic axes default to decade spacing with an optional on/off minor-division switch."
        )

        st.markdown("#### Area Colors")
        color_columns = st.columns(2, gap="large")
        for index, area in enumerate(legend_areas):
            with color_columns[index % 2]:
                area_color_overrides[area] = st.color_picker(
                    area,
                    value=base_area_color_map.get(area, "#4c78a8"),
                    key=f"area_color_{area}",
                )

area_color_map = build_color_map(legend_areas, area_color_overrides)

with left_col:
    clear_highlight = st.button("Clear Highlight", use_container_width=True)
    if clear_highlight:
        st.session_state["highlighted_well"] = None
    for area, wells in legend_groups.items():
        st.markdown(f"**{area}**")
        for well in wells:
            color = area_color_map.get(area, "#4c78a8")
            dot_col, button_col = st.columns([1, 9], gap="small")
            with dot_col:
                st.markdown(
                    f"<div style='color:{color};font-size:18px;line-height:2.1;text-align:center;'>&#9679;</div>",
                    unsafe_allow_html=True,
                )
            with button_col:
                is_selected = st.session_state.get("highlighted_well") == well
                st.button(
                    well,
                    key=f"legend_well_{area}_{well}",
                    use_container_width=True,
                    type="primary" if is_selected else "secondary",
                    on_click=set_highlighted_well,
                    args=(well,),
                )

with right_col:
    with graph_tab:
        chart_tools_col, line_tools_col, chart_note_col = st.columns([1, 1.15, 2.85], gap="medium")
        with chart_tools_col:
            reset_view = st.button("Reset Chart To Settings", use_container_width=True)
        with line_tools_col:
            if st.session_state.get("line_selection_mode"):
                st.button("Cancel Line Tool", use_container_width=True, on_click=clear_line_selection)
            elif len(st.session_state.get("selected_line_points", [])) == 2:
                st.button("Clear Drawn Line", use_container_width=True, on_click=clear_line_selection)
            else:
                st.button("Pick 2 Points", use_container_width=True, on_click=start_line_selection)
        with chart_note_col:
            if st.session_state.get("line_selection_mode"):
                st.caption(
                    "Line tool is active. Click anywhere on the chart area to place two points, "
                    "then the line and its equation will appear below the graph."
                )
            else:
                st.caption(
                    "Zooming or panning with the chart toolbar changes only the temporary chart view. "
                    "Use Reset Chart To Settings to snap the graph back to the values shown in Settings."
                )

        chart_key = "main_plot"
        if reset_view:
            st.session_state["chart_reset_nonce"] = st.session_state.get("chart_reset_nonce", 0) + 1
            st.session_state["highlighted_well"] = None
            st.session_state["selected_line_points"] = []
            st.session_state["line_selection_mode"] = False
        chart_key = f"main_plot_{st.session_state.get('chart_reset_nonce', 0)}"

        validation_errors = []
        validation_errors.extend(validate_axis_settings("X Axis", x_scale, x_start, x_end, x_major))
        validation_errors.extend(validate_axis_settings("Y Axis", y_scale, y_start, y_end, y_major))

        if validation_errors:
            for error in validation_errors:
                st.error(error)
        else:
            figure = build_figure(
                data,
                workbook_title,
                area_color_map,
                st.session_state.get("highlighted_well"),
                st.session_state.get("line_selection_mode", False),
                x_column,
                y_column,
                x_scale,
                y_scale,
                x_start if x_start != x_end else None,
                x_end if x_start != x_end else None,
                x_major or None,
                x_minor or None,
                y_start if y_start != y_end else None,
                y_end if y_start != y_end else None,
                y_major or 10.0,
                y_minor or None,
                st.session_state.get("selected_line_points"),
            )
            st.plotly_chart(
                figure,
                width="stretch",
                key=chart_key,
                on_select=lambda: handle_chart_selection(chart_key),
                selection_mode="points",
            )

            line_equation = build_line_equation(
                st.session_state.get("selected_line_points"),
                x_scale,
                y_scale,
            )
            if line_equation:
                st.markdown(
                    (
                        "<div style='text-align:center; font-size:20px; font-weight:600; "
                        "padding-top:14px;'>Line Equation: "
                        f"{line_equation}</div>"
                    ),
                    unsafe_allow_html=True,
                )

    with preview_tab:
        st.dataframe(data, width="stretch", height=500)
