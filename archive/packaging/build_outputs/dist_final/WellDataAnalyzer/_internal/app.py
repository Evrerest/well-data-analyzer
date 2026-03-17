from __future__ import annotations

from math import ceil, floor, log10
import sys
from pathlib import Path
from typing import Any

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from plotly.colors import qualitative


DEFAULT_SAMPLE_PATH = Path(r"C:\Users\ekost\Downloads\Viking Core Data.xlsx")
KNOWN_HEADERS = {
    1: "Area",
    2: "Well",
    3: "Core",
    4: "Lithology",
    5: "Depth",
    6: "Value_F",
    7: "kmax",
    8: "k90",
    9: "kvert",
    10: "Por",
    11: "GrainDensity",
    16: "Sw",
    17: "BVW",
    22: "Notes",
    23: "Weight_1",
    24: "Weight_2",
    25: "Class",
}
DEFAULT_X = "Por"
DEFAULT_Y = "kmax"
AREA_FALLBACK = "Unassigned Area"


def resolve_default_workbook_path() -> str:
    candidate_paths = []

    bundle_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidate_paths.append(bundle_dir / "sample_data" / "Viking Core Data.xlsx")
    candidate_paths.append(DEFAULT_SAMPLE_PATH)

    for candidate in candidate_paths:
        if candidate.exists():
            return str(candidate)
    return ""


def sanitize_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else None
    return value


def build_headers(worksheet) -> list[str]:
    headers: list[str] = []
    for column in range(1, worksheet.max_column + 1):
        if column in KNOWN_HEADERS:
            headers.append(KNOWN_HEADERS[column])
            continue

        row2 = sanitize_value(worksheet.cell(2, column).value)
        if row2 is not None:
            headers.append(str(row2))
            continue

        headers.append(f"Column_{column}")
    return headers


@st.cache_data(show_spinner=False)
def load_workbook_data(workbook_path: str) -> tuple[str, pd.DataFrame]:
    path = Path(workbook_path)
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    title = str(worksheet["A1"].value or path.stem)
    headers = build_headers(worksheet)

    current_area: str | None = None
    current_well: str | None = None
    records: list[dict[str, Any]] = []

    for row_number in range(3, worksheet.max_row + 1):
        row_values = [sanitize_value(worksheet.cell(row_number, col).value) for col in range(1, worksheet.max_column + 1)]
        if not any(value is not None for value in row_values):
            continue

        if row_values[0] is not None:
            current_area = str(row_values[0])
        if row_values[1] is not None:
            current_well = str(row_values[1])

        record = {header: value for header, value in zip(headers, row_values)}
        record["Area"] = current_area or AREA_FALLBACK
        record["Well"] = current_well or "Unknown Well"
        records.append(record)

    dataframe = pd.DataFrame(records)
    for column in dataframe.columns:
        converted = pd.to_numeric(dataframe[column], errors="coerce")
        non_null_original = dataframe[column].notna().sum()
        non_null_converted = converted.notna().sum()
        if non_null_original > 0 and non_null_original == non_null_converted:
            dataframe[column] = converted

    if "Area" not in dataframe:
        dataframe["Area"] = AREA_FALLBACK
    if "Well" not in dataframe:
        dataframe["Well"] = "Unknown Well"

    dataframe["LegendLabel"] = dataframe["Area"].astype(str) + " | " + dataframe["Well"].astype(str)
    return title, dataframe


def build_color_map(labels: list[str]) -> dict[str, str]:
    palette = qualitative.Safe + qualitative.Bold + qualitative.Dark24
    return {label: palette[index % len(palette)] for index, label in enumerate(labels)}


def nice_linear_tick(raw_step: float | None) -> float | None:
    if raw_step is None or raw_step <= 0:
        return None

    magnitude = 10 ** int(log10(raw_step))
    normalized = raw_step / magnitude
    if normalized <= 1:
        nice = 1
    elif normalized <= 2:
        nice = 2
    elif normalized <= 5:
        nice = 5
    else:
        nice = 10
    return nice * magnitude


def default_tick_step(series: pd.Series) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return None
    data_min = float(numeric.min())
    data_max = float(numeric.max())
    if data_max <= data_min:
        return None
    return nice_linear_tick((data_max - data_min) / 8)


def default_minor_step(major_step: float | None) -> float | None:
    if major_step is None or major_step <= 0:
        return None
    return major_step / 5


def build_log_tick_values(start: float, end: float, minor_per_decade: int) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []

    first_decade = int(log10(start))
    last_decade = int(log10(end)) + 1
    ticks: list[float] = []

    for exponent in range(first_decade, last_decade + 1):
        decade_base = 10**exponent
        for multiplier in range(1, minor_per_decade + 1):
            tick_value = multiplier * decade_base
            if start <= tick_value <= end:
                ticks.append(float(tick_value))
    return ticks


def build_linear_tick_values(start: float, end: float, step: float) -> list[float]:
    if step <= 0 or end <= start:
        return []

    ticks: list[float] = []
    first_index = ceil(start / step)
    current = first_index * step
    limit = end + (step * 1e-9)
    while current <= limit:
        ticks.append(round(current, 12))
        current += step
    return ticks


def build_log_major_ticks(start: float, end: float, major_factor: float) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []
    if major_factor <= 1:
        major_factor = 10

    ticks: list[float] = []
    current_power = ceil(log10(start) / log10(major_factor))
    current = major_factor**current_power
    while current <= end:
        ticks.append(round(float(current), 12))
        current *= major_factor
    return ticks


def build_log_minor_ticks(start: float, end: float, major_factor: float, subdivisions: int) -> list[float]:
    if start <= 0 or end <= 0 or end <= start:
        return []
    if major_factor <= 1 or subdivisions <= 1:
        return []

    ticks: list[float] = []
    current_power = floor(log10(start) / log10(major_factor))
    current_major = major_factor**current_power
    while current_major <= end:
        for index in range(2, subdivisions + 1):
            tick = current_major * index
            if start <= tick <= end and tick < current_major * major_factor:
                ticks.append(round(float(tick), 12))
        current_major *= major_factor
    return ticks


def apply_axis_settings(
    figure: go.Figure,
    axis_name: str,
    title: str,
    scale: str,
    start: float | None,
    end: float | None,
    major_tick: float | None,
    minor_tick: float | None,
) -> None:
    axis_type = "log" if scale == "Logarithmic" else "linear"
    update_method = figure.update_xaxes if axis_name == "x" else figure.update_yaxes

    axis_args: dict[str, Any] = {
        "title_text": title,
        "type": axis_type,
        "showgrid": True,
        "gridcolor": "rgba(120, 120, 120, 0.20)",
        "zeroline": False,
        "tickformat": "g",
    }

    if start is not None and end is not None and end > start:
        if axis_type == "log":
            if start > 0 and end > 0:
                axis_args["range"] = [log10(start), log10(end)]
        else:
            axis_args["range"] = [start, end]

    if axis_type == "linear":
        if start is not None and end is not None and major_tick and major_tick > 0:
            major_ticks = build_linear_tick_values(start, end, major_tick)
            axis_args["tickmode"] = "array"
            axis_args["tickvals"] = major_ticks
            axis_args["ticktext"] = [f"{tick:g}" for tick in major_ticks]

        if start is not None and end is not None and minor_tick and minor_tick > 0:
            minor_ticks = build_linear_tick_values(start, end, minor_tick)
            major_tick_set = set(axis_args.get("tickvals", []))
            minor_only = [tick for tick in minor_ticks if tick not in major_tick_set]
            axis_args["minor"] = {
                "showgrid": True,
                "gridcolor": "rgba(120, 120, 120, 0.10)",
                "tickmode": "array",
                "tickvals": minor_only,
            }
    else:
        if start is not None and end is not None and start > 0 and end > start:
            major_ticks = build_log_major_ticks(start, end, major_tick or 10)
            if major_ticks:
                axis_args["tickmode"] = "array"
                axis_args["tickvals"] = major_ticks
                axis_args["ticktext"] = [f"{tick:g}" for tick in major_ticks]

            if minor_tick and minor_tick > 1:
                subdivisions = max(int(minor_tick), 1)
                axis_args["minor"] = {
                    "showgrid": True,
                    "gridcolor": "rgba(120, 120, 120, 0.10)",
                    "tickmode": "array",
                    "tickvals": build_log_minor_ticks(start, end, major_tick or 10, subdivisions),
                }

    update_method(**axis_args)


def build_figure(
    dataframe: pd.DataFrame,
    x_column: str,
    y_column: str,
    x_scale: str,
    y_scale: str,
    x_start: float | None,
    x_end: float | None,
    x_major: float | None,
    x_minor: float | None,
    y_start: float | None,
    y_end: float | None,
    y_major: float | None,
    y_minor: float | None,
) -> go.Figure:
    plot_data = dataframe[[x_column, y_column, "Area", "Well", "LegendLabel"]].copy()
    plot_data = plot_data.dropna(subset=[x_column, y_column])
    plot_data = plot_data[(plot_data[x_column] != "") & (plot_data[y_column] != "")]

    if x_scale == "Logarithmic":
        plot_data = plot_data[pd.to_numeric(plot_data[x_column], errors="coerce") > 0]
    if y_scale == "Logarithmic":
        plot_data = plot_data[pd.to_numeric(plot_data[y_column], errors="coerce") > 0]

    area_labels = plot_data["Area"].drop_duplicates().tolist()
    color_map = build_color_map(area_labels)

    if plot_data.empty:
        figure = go.Figure()
        figure.update_layout(
            height=720,
            paper_bgcolor="#fbfaf6",
            plot_bgcolor="#fbfaf6",
            annotations=[
                {
                    "text": "No rows match the current axis settings.",
                    "xref": "paper",
                    "yref": "paper",
                    "x": 0.5,
                    "y": 0.5,
                    "showarrow": False,
                    "font": {"size": 18},
                }
            ],
        )
        return figure

    figure = px.scatter(
        plot_data,
        x=x_column,
        y=y_column,
        color="Area",
        symbol="Well",
        color_discrete_map=color_map,
        hover_data={"Area": True, "Well": True},
        labels={x_column: x_column, y_column: y_column, "Area": "Area"},
    )
    figure.update_traces(marker={"size": 10, "line": {"width": 0.5, "color": "#1f2933"}})
    figure.update_layout(
        height=720,
        margin={"l": 20, "r": 20, "t": 20, "b": 20},
        legend_title_text="Area",
        paper_bgcolor="#fbfaf6",
        plot_bgcolor="#fbfaf6",
    )

    apply_axis_settings(figure, "x", x_column, x_scale, x_start, x_end, x_major, x_minor)
    apply_axis_settings(figure, "y", y_column, y_scale, y_start, y_end, y_major, y_minor)
    return figure


def default_axis_value(series: pd.Series, which: str) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return None
    return float(numeric.min() if which == "min" else numeric.max())


def grouped_legend(dataframe: pd.DataFrame) -> dict[str, list[str]]:
    grouped = (
        dataframe[["Area", "Well"]]
        .drop_duplicates()
        .sort_values(["Area", "Well"])
        .groupby("Area")["Well"]
        .apply(list)
        .to_dict()
    )
    return grouped


def axis_input_block(axis_id: str, series: pd.Series, scale: str) -> tuple[float, float, float, float]:
    min_value = default_axis_value(series, "min") or 0.0
    max_value = default_axis_value(series, "max") or 1.0
    if scale == "Logarithmic":
        major_default = 10.0
        minor_default = 9.0
        major_label = "Major Factor"
        minor_label = "Minor Divisions Per Decade"
        major_help = "Use 10 for decade spacing: 0.01, 0.1, 1, 10, 100."
        minor_help = "Use 9 to show intermediate values 2 through 9 within each decade."
    else:
        major_default = default_tick_step(series) or 1.0
        minor_default = default_minor_step(major_default) or 0.2
        major_label = "Major Tick Spacing"
        minor_label = "Minor Tick Spacing"
        major_help = "Exact spacing between labeled major ticks."
        minor_help = "Exact spacing between lighter minor grid lines."

    start = st.number_input(
        "Start",
        value=min_value,
        format="%.6f",
        key=f"{axis_id}_start_{series.name}",
    )
    end = st.number_input(
        "End",
        value=max_value,
        format="%.6f",
        key=f"{axis_id}_end_{series.name}",
    )
    major = st.number_input(
        major_label,
        min_value=float(0.0),
        value=float(major_default),
        format="%.6f",
        key=f"{axis_id}_major_{series.name}",
        help=major_help,
    )
    minor = st.number_input(
        minor_label,
        min_value=float(0.0),
        value=float(minor_default),
        format="%.6f",
        key=f"{axis_id}_minor_{series.name}",
        help=minor_help,
    )
    return start, end, major, minor


st.set_page_config(page_title="Well Data Analyzer", layout="wide")

st.title("Well Data Analyzer")
st.caption("Interactive cross-plotting for core and well data.")

with st.sidebar:
    st.header("Workbook")
    uploaded_file = st.file_uploader("Upload an Excel workbook", type=["xlsx"])
    workbook_path = st.text_input("Or enter a local workbook path", value=resolve_default_workbook_path())

if uploaded_file is not None:
    upload_dir = Path(".streamlit_uploads")
    upload_dir.mkdir(exist_ok=True)
    uploaded_path = upload_dir / uploaded_file.name
    uploaded_path.write_bytes(uploaded_file.getbuffer())
    selected_path = str(uploaded_path.resolve())
else:
    selected_path = workbook_path

path_obj = Path(selected_path)
if not selected_path:
    st.info("Upload a workbook or enter a workbook path to begin.")
    st.stop()

if not path_obj.exists():
    st.error(f"Workbook not found: {path_obj}")
    st.stop()

try:
    workbook_title, data = load_workbook_data(str(path_obj))
except Exception as exc:
    st.exception(exc)
    st.stop()

st.subheader(workbook_title)

numeric_columns = [column for column in data.columns if pd.api.types.is_numeric_dtype(data[column])]
if not numeric_columns:
    st.error("No numeric columns were found in the workbook.")
    st.stop()

default_x_index = numeric_columns.index(DEFAULT_X) if DEFAULT_X in numeric_columns else 0
default_y_index = numeric_columns.index(DEFAULT_Y) if DEFAULT_Y in numeric_columns else min(1, len(numeric_columns) - 1)

legend_groups = grouped_legend(data)
legend_areas = data["Area"].drop_duplicates().tolist()
legend_colors = build_color_map(legend_areas)

left_col, right_col = st.columns([1, 3], gap="large")

with left_col:
    st.markdown("### Legend")
    for area, wells in legend_groups.items():
        st.markdown(f"**{area}**")
        for well in wells:
            color = legend_colors.get(area, "#4c78a8")
            st.markdown(
                f"<span style='color:{color};font-size:18px;'>&#9679;</span> {well}",
                unsafe_allow_html=True,
            )

with right_col:
    graph_tab, settings_tab, preview_tab = st.tabs(["Graph", "Settings", "Data Preview"])

    with settings_tab:
        selector_col, axis_col = st.columns([1, 2], gap="large")

        with selector_col:
            x_column = st.selectbox("X Axis", numeric_columns, index=default_x_index)
            y_column = st.selectbox("Y Axis", numeric_columns, index=default_y_index)

        with axis_col:
            x_settings, y_settings = st.columns(2, gap="large")

            with x_settings:
                st.markdown("#### X Axis Settings")
                x_scale = st.radio("Scale", ["Linear", "Logarithmic"], key="x_scale")
                x_start, x_end, x_major, x_minor = axis_input_block("x", data[x_column], x_scale)

            with y_settings:
                st.markdown("#### Y Axis Settings")
                y_scale = st.radio("Scale", ["Linear", "Logarithmic"], key="y_scale")
                y_start, y_end, y_major, y_minor = axis_input_block("y", data[y_column], y_scale)

        st.caption(
            "Linear axes use exact spacing values. "
            "Logarithmic axes use a major factor and minor divisions per decade."
        )

    with graph_tab:
        chart_tools_col, chart_note_col = st.columns([1, 3], gap="large")
        with chart_tools_col:
            reset_view = st.button("Reset Chart To Settings", use_container_width=True)
        with chart_note_col:
            st.caption(
                "Zooming or panning with the chart toolbar changes only the temporary chart view. "
                "Use Reset Chart To Settings to snap the graph back to the values shown in Settings."
            )

        chart_key = "main_plot"
        if reset_view:
            st.session_state["chart_reset_nonce"] = st.session_state.get("chart_reset_nonce", 0) + 1
        chart_key = f"main_plot_{st.session_state.get('chart_reset_nonce', 0)}"

        figure = build_figure(
            data,
            x_column,
            y_column,
            x_scale,
            y_scale,
            x_start if x_start != x_end else None,
            x_end if x_start != x_end else None,
            x_major or None,
            x_minor or None,
            y_start if y_start != y_end else None,
            y_end if y_start != y_end else None,
            y_major or None,
            y_minor or None,
        )
        st.plotly_chart(figure, width="stretch", key=chart_key)

    with preview_tab:
        st.dataframe(data, width="stretch", height=500)
